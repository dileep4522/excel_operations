import openpyxl
from openpyxl import Workbook,load_workbook
path=r"C:\Users\DILEEP KUMAR\Desktop\dileepnumber.xlsx"
outputfile=r"C:\Users\DILEEP KUMAR\Desktop\updates.xlsx"


wb=openpyxl.load_workbook(path)
sheet1=wb.active

column1='B'
column2='D'

new_wb=Workbook()
sheet2=new_wb.active

for row in range(2,sheet1.max_row + 1):
    obj1=sheet1[f'{column1}{row}'].value
    obj2=sheet1[f'{column2}{row}'].value
    sheet2[f'B{row}']=obj1 + obj2

new_wb.save(outputfile)
print(f'data has been saved{outputfile}')
