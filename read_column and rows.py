import openpyxl

from openpyxl import workbook,load_workbook

path=r"C:\Users\DILEEP KUMAR\Desktop\dileepnumber.xlsx"

wb=openpyxl.load_workbook(path)

sheet=wb.active

colvalue=sheet.max_column

rowvalue=sheet.max_row

for i in range(1,colvalue+1):
    obj1=sheet.cell(row=1,column=i)

    print(obj1.value)

for j in range(1,rowvalue+1):
    obj2=sheet.cell(row=j,column=11)

    print((obj2.value))
