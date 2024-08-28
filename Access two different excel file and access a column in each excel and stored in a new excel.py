mport openpyxl
from openpyxl import Workbook,load_workbook
path=r"C:\Users\DILEEP KUMAR\Desktop\dileepnumber.xlsx"
path2=r"F:\idle\excel\matches.xlsx"
outputfile=r"C:\Users\DILEEP KUMAR\Desktop\updates.xlsx"


if not os.path.exists(path):
    print(f"file not found {path}")
    exit()
if not os.path.exists(path2):
    print(f"file not found{path2}")

wb=openpyxl.load_workbook(path)
sheet1=wb.active
wb2=openpyxl.load_workbook(path2)
sheet2=wb2.active

select_col1=sheet1['B']
select_col2=sheet2['C']

new_wb=openpyxl.load_workbook(outputfile)
new_sheet2=new_wb.active

for i,cell in enumerate(select_col1,start=1):
    new_sheet2.cell(row=i,column=1,value=cell.value)

for j,cell in enumerate(select_col2,start=1):
    new_sheet2.cell(row=j,column=2,value=cell.value)
new_wb.save(outputfile)
