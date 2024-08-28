import openpyxl
from openpyxl import Workbook,load_workbook
path=r"C:\Users\DILEEP KUMAR\Desktop\dileepnumber.xlsx"
outputfile=r"C:\Users\DILEEP KUMAR\Desktop\updates.xlsx"
wb=openpyxl.load_workbook(path)
sheet1=wb.active

column1='A'
column2='B'

start_row=int(input("enter start position"))
end_row=int(input("enter end position"))

data_column1=[]
data_column2=[]


for row in range(start_row,end_row+1):
    obj1=sheet1[f'{column1}{row}'].value
    obj2=sheet1[f'{column2}{row}'].value
    data_column1.append(obj1)
    data_column2.append(obj2)

new_wb = Workbook()
sheet2 = new_wb.active

for index, (val1, val2) in enumerate(zip(data_column1, data_column2), start=1):
    sheet2[f'A{index}'] = val1
    sheet2[f'B{index}'] = val2
new_wb.save(outputfile)
print(f'data has been saved in {outputfile}')

