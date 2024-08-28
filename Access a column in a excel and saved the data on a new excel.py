import openpyxl
from openpyxl import Workbook,load_workbook
path=r"C:\Users\DILEEP KUMAR\Desktop\dileepnumber.xlsx"
wb=openpyxl.load_workbook(path)
sheet=wb.active
select_col='B'
saved_column=[]


for cell in sheet[select_col]:
    saved_column.append(cell.value)

new_wb=Workbook()

new_sheet=new_wb.active


for row_index,value in enumerate(saved_column,start=1):
    new_sheet[f'B{row_index}']=value

outputfile=r"C:\Users\DILEEP KUMAR\Desktop\updates.xlsx"
new_wb.save(filename=outputfile)
print("saved")
print(f"columns {select_col}data has been saven {outputfile}")
