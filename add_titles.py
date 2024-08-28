import openpyxl
from openpyxl import Workbook,load_workbook
path=r"C:\Users\DILEEP KUMAR\Desktop\updates.xlsx"
wb=openpyxl.load_workbook(path)
sheet=wb.active
titles=['total']
columnname,titles in enumerate(titles,2):
    cell = sheet.cell(row=1,column=columnname)
    cell.value=titles
wb.save(r"C:\Users\DILEEP KUMAR\Desktop\updates.xlsx")
